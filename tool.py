#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
vn_parser_gui_with_config.py – Visual‑Novel scenario → Excel converter (configurable)

Updates:
- Menu bar: File, Language (English/Tiếng Việt), Help
- Tabs: File processing, Insert again (placeholder), Alice tool
- Auto-save/load config: remembers last input/output and rules
- Improved UI with inline rule editing and statistics

Run:
> pip install openpyxl
> python vn_parser_gui_with_config.py

© ChatGPT, 2025‑06‑28
"""

import os
import re
import json
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
import sys

CONFIG_FILE = 'vn_config.json'

# Language dictionaries
LANG = {
    'en': {
        'title': 'Tsumamigui 3 Tool',
        'file_menu': 'File',
        'lang_menu': 'Language',
        'help_menu': 'Help',
        'help_title': 'How to use',
        'help_text': """TSUMAMIGUI 3 TRANSLATION WORKFLOW:

=== TAB 1: FILE PROCESSING ===
1. Choose input TXT file (scenario file from game)
2. Choose output Excel file (for translation work)
3. Add dialogue delimiter rules with start/end chars (e.g. 「/」, 『/』, （/）)
4. Use +/-/↑/↓ to manage rule priority (top = highest priority)
5. Click Convert to export dialogues to Excel for translation

=== TAB 2: INSERT AGAIN ===
1. Choose Excel file (with completed translations in Translate column)
2. Choose TXT file (original scenario file to modify)
3. Configure max characters per line and character mapping
4. Click Insert to apply translations back to TXT file
   • Empty translations = skip (keep original commented ;m[])
   • "null" translations = uncomment but set empty m[] = ""
   • Real translations = uncomment and insert translated text

=== TAB 3: ALICE TOOL ===
1. Choose Ain file (game's script file)
2. Choose TXT file (with translations applied)
3. Choose Output path (where to save new Ain file)
4. Click Pack Ain File to compile into game format

=== COMPLETE WORKFLOW ===
TXT (original) → Excel (translate) → TXT (translated) → AIN (game ready)

All settings auto-save and restore on restart.""",
        'tab_file': 'File processing',
        'tab_insert': 'Insert again',
        'tab_alice': 'Alice tool',
        'btn_txt': 'TXT…',
        'btn_excel': 'Excel…',
        'lbl_rules': 'Dialogue delimiter rules (top priority first):',
        'col_start': 'Start',
        'col_end': 'End',
        'btn_convert': 'Convert',
        'btn_add': 'Add Rule',
        'lbl_start': 'Start:',
        'lbl_end': 'End:',
        'btn_ok': 'OK',
        'msg_no_files': 'Choose TXT and Excel files.',
        'msg_no_rules': 'Add at least one delimiter rule.',
        'msg_end_required': 'End delimiter is required!',
        'msg_finished': 'Export completed.',
        'msg_error': 'Error',
        'lbl_stats': 'Statistics:',
        'lbl_total_lines': 'Total lines:',
        'lbl_dialogue_segments': 'Dialogue segments:',
        'lbl_progress': 'Progress:'
    },
    'vi': {
        'title': 'Tsumamigui 3 Tool',
        'file_menu': 'Tệp',
        'lang_menu': 'Ngôn ngữ',
        'help_menu': 'Trợ giúp',
        'help_title': 'Cách sử dụng',
        'help_text': """TSUMAMIGUI 3 TRANSLATION WORKFLOW:

=== TAB 1: FILE PROCESSING ===
1. Chọn tệp TXT đầu vào (file cảnh báo từ game)
2. Chọn tệp Excel đầu ra (cho công việc dịch thuật)
3. Thêm quy tắc hội thoại với ký tự bắt đầu và kết thúc (ví dụ:「 / 」, 『/』, （/）)
4. Dùng + / - / ↑ / ↓ để quản lý độ ưu tiên quy tắc (top = độ ưu tiên cao nhất)
5. Nhấn Chuyển đổi để xuất hội thoại ra Excel cho dịch thuật

=== TAB 2: CHÈN LẠI ===
1. Chọn tệp Excel (với dịch thuật hoàn thành trong cột Translate)
2. Chọn tệp TXT (file cảnh báo gốc để sửa đổi)
3. Cấu hình số ký tự tối đa trên mỗi dòng và ánh xạ ký tự
4. Nhấn Chèn để áp dụng dịch thuật trở lại tệp TXT
   • Dịch thuật trống = bỏ qua (giữ nguyên comment ;m[])
   • "null" dịch thuật = bỏ comment nhưng đặt m[] = ""
   • Dịch thuật thực = bỏ comment và chèn văn bản đã dịch

=== TAB 3: ALICE TOOL ===
1. Chọn tệp Ain (file script của game)
2. Chọn tệp TXT (với dịch thuật đã áp dụng)
3. Chọn Đường dẫn đầu ra (nơi để lưu tệp Ain mới)
4. Nhấn Pack Ain File để biên dịch thành định dạng game

=== HOÀN THÀNH CÔNG VIỆC ===
TXT (gốc) → Excel (dịch) → TXT (đã dịch) → AIN (game sẵn sàng)

Tất cả các cài đặt tự động lưu và khôi phục lại khi khởi động lại.""",
        'tab_file': 'Xử lý tệp',
        'tab_insert': 'Chèn lại',
        'tab_alice': 'Alice tool',
        'btn_txt': 'TXT…',
        'btn_excel': 'Excel…',
        'lbl_rules': 'Quy tắc phân định hội thoại (ưu tiên từ trên xuống):',
        'col_start': 'Bắt đầu',
        'col_end': 'Kết thúc',
        'btn_convert': 'Chuyển đổi',
        'btn_add': 'Thêm quy tắc',
        'lbl_start': 'Bắt đầu:',
        'lbl_end': 'Kết thúc:',
        'btn_ok': 'OK',
        'msg_no_files': 'Hãy chọn tệp TXT và Excel.',
        'msg_no_rules': 'Thêm ít nhất một quy tắc phân định.',
        'msg_end_required': 'Ký tự kết thúc là bắt buộc!',
        'msg_finished': 'Xuất hoàn tất.',
        'msg_error': 'Lỗi',
        'lbl_stats': 'Thống kê:',
        'lbl_total_lines': 'Tổng số dòng:',
        'lbl_dialogue_segments': 'Đoạn hội thoại:',
        'lbl_progress': 'Tiến độ:'
    }
}

re_m = re.compile(r'^;m\[(\d+)]\s*=\s*"(.*)"')
re_s = re.compile(r'^;s\[(\d+)]\s*=\s*"(.*)"')
re_has_letter = re.compile(r'\D')

def speaker_finder(line: str):
    m = re_s.match(line)
    if m and re_has_letter.search(m.group(2)):
        return m.group(2)
    return None

def parse_stream(txt_path: str, rules: list[dict], q_msg: queue.Queue):
    file_size = os.path.getsize(txt_path)
    read_bytes = 0
    recent_speakers = []  # List to store recent s[] entries

    with open(txt_path, encoding='utf-8', errors='ignore') as fh:
        buf, cur_rule = [], None
        start_tag, start_line, line_idx = None, None, -1

        def find_matching_rule(text):
            """Find the best matching rule for given text, prioritizing non-empty start rules"""
            # First pass: check rules with non-empty start
            for r in rules:
                if r['start'] != '' and text.startswith(r['start']):
                    return r
            
            # Second pass: check rules with empty start
            for r in rules:
                if r['start'] == '':
                    return r
            
            # Fallback: create a rule based on last character
            return {'start': '', 'end': text[-1:]}

        def should_start_new_segment(text, current_rule):
            """Check if we should start a new segment for this text"""
            if not buf:
                return True
            
            # Check if this text matches a rule with explicit start delimiter
            for r in rules:
                if r['start'] != '' and text.startswith(r['start']):
                    return True
            
            return False

        def get_speaker_for_segment():
            """Get the first speaker with letters from recent_speakers list"""
            for speaker in recent_speakers:
                if re_has_letter.search(speaker):
                    return speaker
            return ''

        while True:
            line = fh.readline()
            if not line:
                # Flush remaining buffer at end of file
                if buf:
                    rng = f"{start_tag}" if len(buf) == 1 else f"{start_tag}-{buf[-1][0]}"
                    spk = get_speaker_for_segment()
                    yield rng, spk, ''.join(t for _, t in buf)
                break

            line_idx += 1
            read_bytes += len(line)

            spk_tmp = speaker_finder(line)
            if spk_tmp is not None:
                # If we encounter s[] line while having buffered m[] lines, flush the buffer
                if buf:
                    rng = f"{start_tag}" if len(buf) == 1 else f"{start_tag}-{buf[-1][0]}"
                    spk = get_speaker_for_segment()
                    yield rng, spk, ''.join(t for _, t in buf)
                    buf.clear()
                    cur_rule = None
                    start_tag = None
                    recent_speakers.clear()  # Clear speakers after flushing
                
                # Add speaker to recent list
                recent_speakers.append(spk_tmp)
                continue

            m = re_m.match(line)
            if not m:
                continue

            tag_num, m_text = int(m.group(1)), m.group(2)

            # Check if we should start a new segment
            if should_start_new_segment(m_text, cur_rule):
                # Flush current buffer if exists
                if buf:
                    rng = f"{start_tag}" if len(buf) == 1 else f"{start_tag}-{buf[-1][0]}"
                    spk = get_speaker_for_segment()
                    yield rng, spk, ''.join(t for _, t in buf)
                    buf.clear()
                
                # Start new segment
                start_tag, start_line = tag_num, line_idx
                cur_rule = find_matching_rule(m_text)

            buf.append((tag_num, m_text))

            # Check if current dialogue segment is complete
            if cur_rule and m_text.endswith(cur_rule['end']):
                rng = f"{start_tag}" if len(buf) == 1 else f"{start_tag}-{buf[-1][0]}"
                spk = get_speaker_for_segment()
                yield rng, spk, ''.join(t for _, t in buf)
                buf.clear()
                cur_rule = None
                start_tag = None
                recent_speakers.clear()  # Clear speakers after completing segment

            if (read_bytes & ((1 << 20) - 1)) < len(line):
                q_msg.put(('progress', read_bytes / file_size))


def save_config(txt_path, out_path, rules, insert_config=None):
    config = {
        'txt_path': txt_path,
        'out_path': out_path,
        'rules': rules
    }
    if insert_config:
        config['insert_config'] = insert_config
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def save_insert_config(insert_input_path, insert_output_path, max_chars, vir_chars, phy_chars, show_vir, show_phy):
    """Save insert tab configuration"""
    # Load existing config first
    existing_config = load_config() or {}
    
    insert_config = {
        'insert_input_path': insert_input_path,
        'insert_output_path': insert_output_path,
        'max_chars': max_chars,
        'vir_chars': vir_chars,
        'phy_chars': phy_chars,
        'show_vir': show_vir,
        'show_phy': show_phy
    }
    
    existing_config['insert_config'] = insert_config
    
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(existing_config, f, ensure_ascii=False, indent=2)

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return None
    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.language = 'en'
        self.txt_path = ''
        self.out_path = ''
        self.queue = queue.Queue()
        self.stats = {'total_lines': 0, 'dialogue_segments': 0}
        
        self.setup_ui()
        self.load_saved_config()
        self.after(100, self.listen_queue)

    def setup_ui(self):
        self.title(LANG[self.language]['title'])
        self.geometry('700x650')  # Increased height to show Convert button
        self.resizable(False, False)
        
        self.create_menu()
        self.create_tabs()

    def create_menu(self):
        menubar = tk.Menu(self)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label='(empty)')
        menubar.add_cascade(label=LANG[self.language]['file_menu'], menu=file_menu)

        lang_menu = tk.Menu(menubar, tearoff=0)
        lang_menu.add_command(label='English', command=lambda: self.set_language('en'))
        lang_menu.add_command(label='Tiếng Việt', command=lambda: self.set_language('vi'))
        menubar.add_cascade(label=LANG[self.language]['lang_menu'], menu=lang_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label=LANG[self.language]['help_title'], command=self.show_help)
        menubar.add_cascade(label=LANG[self.language]['help_menu'], menu=help_menu)

        self.config(menu=menubar)

    def set_language(self, lang):
        self.language = lang
        self.refresh_ui()
        messagebox.showinfo('Language', f'Language set to {"English" if lang == "en" else "Tiếng Việt"}.')

    def refresh_ui(self):
        # Destroy and recreate UI with new language
        for widget in self.winfo_children():
            widget.destroy()
        self.setup_ui()
        self.load_saved_config()

    def show_help(self):
        messagebox.showinfo(LANG[self.language]['help_title'], LANG[self.language]['help_text'])

    def create_tabs(self):
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill='both', expand=True, padx=10, pady=5)

        self.tab_file = ttk.Frame(self.nb)
        self.nb.add(self.tab_file, text=LANG[self.language]['tab_file'])

        self.tab_insert = ttk.Frame(self.nb)
        self.nb.add(self.tab_insert, text=LANG[self.language]['tab_insert'])

        self.tab_alice = ttk.Frame(self.nb)
        self.nb.add(self.tab_alice, text="Alice tool")

        self.create_file_processing_widgets()
        self.create_insert_widgets()
        self.create_alice_widgets()

    def create_file_processing_widgets(self):
        # File selection section
        file_frame = ttk.LabelFrame(self.tab_file, text="File Selection", padding=10)
        file_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(file_frame, text=LANG[self.language]['btn_txt'], width=12, command=self.choose_input)\
            .grid(row=0, column=0, padx=5, sticky='w')
        self.lbl_in = ttk.Label(file_frame, text='–', foreground='gray')
        self.lbl_in.grid(row=0, column=1, sticky='w', padx=(10, 0))

        ttk.Button(file_frame, text=LANG[self.language]['btn_excel'], width=12, command=self.choose_output)\
            .grid(row=1, column=0, padx=5, pady=6, sticky='w')
        self.lbl_out = ttk.Label(file_frame, text='–', foreground='gray')
        self.lbl_out.grid(row=1, column=1, sticky='w', padx=(10, 0))

        # Rules section
        rules_frame = ttk.LabelFrame(self.tab_file, text=LANG[self.language]['lbl_rules'], padding=10)
        rules_frame.pack(fill='both', expand=True, padx=10, pady=5)

        # Inline rule addition
        add_frame = ttk.Frame(rules_frame)
        add_frame.pack(fill='x', pady=(0, 10))

        ttk.Label(add_frame, text=LANG[self.language]['lbl_start']).grid(row=0, column=0, padx=(0, 5))
        self.entry_start = ttk.Entry(add_frame, width=12)
        self.entry_start.grid(row=0, column=1, padx=5)
        
        ttk.Label(add_frame, text=LANG[self.language]['lbl_end']).grid(row=0, column=2, padx=(10, 5))
        self.entry_end = ttk.Entry(add_frame, width=12)
        self.entry_end.grid(row=0, column=3, padx=5)
        
        ttk.Button(add_frame, text=LANG[self.language]['btn_add'], command=self.add_rule_inline)\
            .grid(row=0, column=4, padx=(10, 0))

        # Rules tree
        tree_frame = ttk.Frame(rules_frame)
        tree_frame.pack(fill='both', expand=True)

        self.tree = ttk.Treeview(tree_frame, columns=('start', 'end'), show='headings', height=8)
        self.tree.heading('start', text=LANG[self.language]['col_start'])
        self.tree.heading('end', text=LANG[self.language]['col_end'])
        self.tree.column('start', width=150, anchor='center')
        self.tree.column('end', width=150, anchor='center')
        
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Tree control buttons
        btn_frame = ttk.Frame(rules_frame)
        btn_frame.pack(pady=(10, 0))
        
        ttk.Button(btn_frame, text='–', width=4, command=self.del_rule).grid(row=0, column=0, padx=2)
        ttk.Button(btn_frame, text='↑', width=4, command=lambda: self.move_rule(-1)).grid(row=0, column=1, padx=2)
        ttk.Button(btn_frame, text='↓', width=4, command=lambda: self.move_rule(1)).grid(row=0, column=2, padx=2)

        # Convert and statistics section
        bottom_frame = ttk.Frame(self.tab_file)
        bottom_frame.pack(fill='x', padx=10, pady=10)

        # Statistics
        stats_frame = ttk.LabelFrame(bottom_frame, text=LANG[self.language]['lbl_stats'], padding=5)
        stats_frame.pack(fill='x', pady=(0, 10))

        self.lbl_total = ttk.Label(stats_frame, text=f"{LANG[self.language]['lbl_total_lines']} 0")
        self.lbl_total.pack(anchor='w')
        
        self.lbl_segments = ttk.Label(stats_frame, text=f"{LANG[self.language]['lbl_dialogue_segments']} 0")
        self.lbl_segments.pack(anchor='w')

        # Convert button and progress - Make it more visible
        convert_frame = ttk.Frame(bottom_frame)
        convert_frame.pack(fill='x', pady=10)
        
        convert_btn = ttk.Button(convert_frame, text=LANG[self.language]['btn_convert'], width=25, 
                  command=self.run_thread)
        convert_btn.pack(pady=10)

        ttk.Label(bottom_frame, text=LANG[self.language]['lbl_progress']).pack(anchor='w')
        self.pb = ttk.Progressbar(bottom_frame, orient='horizontal', length=500, mode='determinate')
        self.pb.pack(pady=(5, 0))

        # Bind Enter key to add rule
        self.entry_end.bind('<Return>', lambda e: self.add_rule_inline())

    def choose_input(self):
        p = filedialog.askopenfilename(filetypes=[('TXT files', '*.txt')])
        if p:
            self.txt_path = p
            filename = os.path.basename(p)
            self.lbl_in.configure(text=filename, foreground='black')
            self.update_file_stats()

    def choose_output(self):
        p = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])
        if p:
            self.out_path = p
            filename = os.path.basename(p)
            self.lbl_out.configure(text=filename, foreground='black')

    def update_file_stats(self):
        if self.txt_path and os.path.exists(self.txt_path):
            try:
                with open(self.txt_path, 'r', encoding='utf-8', errors='ignore') as f:
                    total_lines = sum(1 for _ in f)
                    self.stats['total_lines'] = total_lines
                    self.lbl_total.configure(text=f"{LANG[self.language]['lbl_total_lines']} {total_lines}")
            except:
                pass

    def add_rule_inline(self):
        start = self.entry_start.get()
        end = self.entry_end.get()
        
        if not end:
            messagebox.showwarning('Rule', LANG[self.language]['msg_end_required'])
            return
        
        self.tree.insert('', 'end', values=(start, end))
        self.entry_start.delete(0, 'end')
        self.entry_end.delete(0, 'end')
        self.entry_start.focus()

    def del_rule(self):
        sel = self.tree.selection()
        if sel:
            self.tree.delete(sel[0])

    def move_rule(self, delta):
        sel = self.tree.selection()
        if not sel:
            return
        item = sel[0]
        idx = self.tree.index(item)
        new_idx = idx + delta
        children = self.tree.get_children()
        if 0 <= new_idx < len(children):
            vals = self.tree.item(item, 'values')
            self.tree.delete(item)
            self.tree.insert('', new_idx, values=vals)
            self.tree.selection_set(self.tree.get_children()[new_idx])

    def collect_rules(self):
        return [{'start': s, 'end': e} for s, e in (self.tree.item(i, 'values') for i in self.tree.get_children())]

    def run_thread(self):
        if not self.txt_path or not self.out_path:
            messagebox.showwarning('Input', LANG[self.language]['msg_no_files'])
            return
        if not self.tree.get_children():
            messagebox.showwarning('Rules', LANG[self.language]['msg_no_rules'])
            return

        self.pb['value'] = 0
        self.stats['dialogue_segments'] = 0
        rules = self.collect_rules()
        save_config(self.txt_path, self.out_path, rules)
        self.toggle_widgets(disable=True)
        threading.Thread(target=self.worker, args=(rules,), daemon=True).start()

    def worker(self, rules):
        try:
            wb = Workbook(write_only=True)
            ws = wb.create_sheet('Dialogues')
            ws.append(['Range', 'Speaker', 'Dialogue', 'Translate'])
            
            segment_count = 0
            for rng, spk, txt in parse_stream(self.txt_path, rules, self.queue):
                ws.append([rng, spk, txt, ''])
                segment_count += 1
                
            wb.save(self.out_path)
            self.queue.put(('segments', segment_count))
            self.queue.put(('done', LANG[self.language]['msg_finished']))
        except Exception as exc:
            self.queue.put(('error', str(exc)))

    def listen_queue(self):
        try:
            while True:
                msg, data = self.queue.get_nowait()
                if msg == 'progress':
                    self.pb['value'] = data * 100
                elif msg == 'segments':
                    self.stats['dialogue_segments'] = data
                    self.lbl_segments.configure(text=f"{LANG[self.language]['lbl_dialogue_segments']} {data}")
                elif msg == 'done':
                    self.pb['value'] = 100
                    messagebox.showinfo('Finished', data)
                    self.toggle_widgets(disable=False)
                elif msg == 'error':
                    messagebox.showerror(LANG[self.language]['msg_error'], data)
                    self.toggle_widgets(disable=False)
                elif msg == 'insert_progress':
                    self.pb_insert['value'] = data * 100
                elif msg == 'insert_done':
                    self.pb_insert['value'] = 100
                    messagebox.showinfo('Insert Finished', data)
                elif msg == 'insert_error':
                    messagebox.showerror('Insert Error', data)
                elif msg == 'alice_done':
                    self.pb_alice.stop()
                    messagebox.showinfo('Alice Pack Finished', data)
                elif msg == 'alice_error':
                    self.pb_alice.stop()
                    messagebox.showerror('Alice Pack Error', data)
        except queue.Empty:
            pass
        self.after(100, self.listen_queue)

    def toggle_widgets(self, disable):
        state = ['disabled'] if disable else ['!disabled']
        for child in self.winfo_children():
            if isinstance(child, ttk.Button):
                child.state(state)

    def load_saved_config(self):
        cfg = load_config()
        if not cfg:
            return
        self.txt_path = cfg.get('txt_path', '')
        self.out_path = cfg.get('out_path', '')
        
        if self.txt_path:
            self.lbl_in.configure(text=os.path.basename(self.txt_path), foreground='black')
            self.update_file_stats()
        if self.out_path:
            self.lbl_out.configure(text=os.path.basename(self.out_path), foreground='black')
            
        for rule in cfg.get('rules', []):
            self.tree.insert('', 'end', values=(rule['start'], rule['end']))
        
        # Load insert config if exists
        insert_config = cfg.get('insert_config', {})
        if insert_config:
            # Load insert file paths
            self.insert_input_path = insert_config.get('insert_input_path', '')
            self.insert_output_path = insert_config.get('insert_output_path', '')
            
            if self.insert_input_path and hasattr(self, 'lbl_insert_in'):
                self.lbl_insert_in.configure(text=os.path.basename(self.insert_input_path), foreground='black')
            if self.insert_output_path and hasattr(self, 'lbl_insert_out'):
                self.lbl_insert_out.configure(text=os.path.basename(self.insert_output_path), foreground='black')
            
            # Load insert settings
            if hasattr(self, 'max_chars_var'):
                self.max_chars_var.set(insert_config.get('max_chars', 50))
                self.vir_chars_var.set(insert_config.get('vir_chars', self.vir_chars_var.get()))
                self.phy_chars_var.set(insert_config.get('phy_chars', self.phy_chars_var.get()))
                self.show_vir.set(insert_config.get('show_vir', False))
                self.show_phy.set(insert_config.get('show_phy', False))
                
                # Update UI state based on loaded settings
                self.toggle_vir_chars()
                self.toggle_phy_chars()

        # Load alice config if exists
        alice_config = cfg.get('alice_config', {})
        if alice_config:
            # Load alice file paths
            self.ain_file_path = alice_config.get('ain_file_path', '')
            self.txt_file_path = alice_config.get('txt_file_path', '')
            self.output_ain_path = alice_config.get('output_ain_path', '')
            
            if self.ain_file_path and hasattr(self, 'lbl_ain_file'):
                self.lbl_ain_file.configure(text=os.path.basename(self.ain_file_path), foreground='black')
            if self.txt_file_path and hasattr(self, 'lbl_txt_file'):
                self.lbl_txt_file.configure(text=os.path.basename(self.txt_file_path), foreground='black')
            if self.output_ain_path and hasattr(self, 'lbl_output_ain'):
                self.lbl_output_ain.configure(text=os.path.basename(self.output_ain_path), foreground='black')

    def create_insert_widgets(self):
        # File selection section for Insert tab
        insert_file_frame = ttk.LabelFrame(self.tab_insert, text="File Selection", padding=10)
        insert_file_frame.pack(fill='x', padx=10, pady=5)

        ttk.Button(insert_file_frame, text="Excel…", width=12, command=self.choose_insert_input)\
            .grid(row=0, column=0, padx=5, sticky='w')
        self.lbl_insert_in = ttk.Label(insert_file_frame, text='–', foreground='gray')
        self.lbl_insert_in.grid(row=0, column=1, sticky='w', padx=(10, 0))

        ttk.Button(insert_file_frame, text="TXT…", width=12, command=self.choose_insert_output)\
            .grid(row=1, column=0, padx=5, pady=6, sticky='w')
        self.lbl_insert_out = ttk.Label(insert_file_frame, text='–', foreground='gray')
        self.lbl_insert_out.grid(row=1, column=1, sticky='w', padx=(10, 0))

        # Configuration section
        config_frame = ttk.LabelFrame(self.tab_insert, text="Configuration", padding=10)
        config_frame.pack(fill='x', padx=10, pady=5)

        # Max characters
        ttk.Label(config_frame, text="Max characters:").grid(row=0, column=0, padx=(0, 5), sticky='w')
        self.max_chars_var = tk.IntVar(value=50)
        max_chars_entry = ttk.Entry(config_frame, textvariable=self.max_chars_var, width=10)
        max_chars_entry.grid(row=0, column=1, padx=5, sticky='w')
        max_chars_entry.bind('<FocusOut>', lambda e: self.save_insert_config_now())
        max_chars_entry.bind('<Return>', lambda e: self.save_insert_config_now())

        # Virtual Characters with toggle
        ttk.Label(config_frame, text="Virtual Characters:").grid(row=1, column=0, padx=(0, 5), sticky='w', pady=(10, 0))
        vir_frame = ttk.Frame(config_frame)
        vir_frame.grid(row=1, column=1, columnspan=2, sticky='ew', pady=(10, 0))
        
        self.show_vir = tk.BooleanVar()
        self.vir_check = ttk.Checkbutton(vir_frame, text="Show/Edit", variable=self.show_vir, command=self.toggle_vir_chars_and_save)
        self.vir_check.pack(side='left')
        
        self.vir_chars_var = tk.StringVar(value="áàảãạắằẳẵặấầẩẫậéèẻẽẹếềểễệíìỉĩịóòỏõọốồổỗộớờởỡợúùủũụứừửữựýỳỷỹỵđÁÀẢÃẠẮẰẲẴẶÉÈẺẼẸẾỀỂỄỆÍÌỈĨỊÓÒỎÕỌỐỒỔỖỘỚỜỞỠỢÚÙỦŨỤỨỪỬỮỰÝỲỶỸỴĐôâăơưÔÂĂƠƯêÊ")
        self.vir_entry = ttk.Entry(vir_frame, textvariable=self.vir_chars_var, width=60, state='disabled')
        self.vir_entry.pack(side='left', padx=(10, 0), fill='x', expand=True)
        self.vir_entry.bind('<FocusOut>', lambda e: self.save_insert_config_now())
        self.vir_entry.bind('<KeyRelease>', lambda e: self.after(500, self.save_insert_config_now))  # Delay save while typing

        # Physical Characters with toggle
        ttk.Label(config_frame, text="Physical Characters:").grid(row=2, column=0, padx=(0, 5), sticky='w', pady=(10, 0))
        phy_frame = ttk.Frame(config_frame)
        phy_frame.grid(row=2, column=1, columnspan=2, sticky='ew', pady=(10, 0))
        
        self.show_phy = tk.BooleanVar()
        self.phy_check = ttk.Checkbutton(phy_frame, text="Show/Edit", variable=self.show_phy, command=self.toggle_phy_chars_and_save)
        self.phy_check.pack(side='left')
        
        self.phy_chars_var = tk.StringVar(value="｡ュョ､･ｦｧｨｩｪｫｬｭｮｯｰｱｲｳｴｵｶｷｸｹｺｻｼｽｾｿﾀﾁﾂﾃﾄﾅﾆﾇﾈﾉﾊﾋﾌﾍﾎﾏﾐﾑﾒﾓﾔﾕﾖﾗﾘﾙﾚﾛﾜﾝアイウエオカキクケコカキクケコサシスセソタチツテトナニヌネノハヒフヘホマミムメモヤユヨラリルレワヲンァィゥェォャあいうえおかきくけこさし")
        self.phy_entry = ttk.Entry(phy_frame, textvariable=self.phy_chars_var, width=60, state='disabled')
        self.phy_entry.pack(side='left', padx=(10, 0), fill='x', expand=True)
        self.phy_entry.bind('<FocusOut>', lambda e: self.save_insert_config_now())
        self.phy_entry.bind('<KeyRelease>', lambda e: self.after(500, self.save_insert_config_now))  # Delay save while typing

        config_frame.columnconfigure(1, weight=1)

        # Insert section
        insert_action_frame = ttk.Frame(self.tab_insert)
        insert_action_frame.pack(fill='x', padx=10, pady=20)

        # Insert button
        insert_btn = ttk.Button(insert_action_frame, text="Insert", width=25, command=self.run_insert)
        insert_btn.pack(pady=10)

        # Progress for insert
        ttk.Label(insert_action_frame, text="Insert Progress:").pack(anchor='w')
        self.pb_insert = ttk.Progressbar(insert_action_frame, orient='horizontal', length=500, mode='determinate')
        self.pb_insert.pack(pady=(5, 0))

        # Initialize paths
        self.insert_input_path = ''
        self.insert_output_path = ''

    def toggle_vir_chars(self):
        if self.show_vir.get():
            self.vir_entry.config(state='normal')
        else:
            self.vir_entry.config(state='disabled')

    def toggle_phy_chars(self):
        if self.show_phy.get():
            self.phy_entry.config(state='normal')
        else:
            self.phy_entry.config(state='disabled')

    def toggle_vir_chars_and_save(self):
        self.toggle_vir_chars()
        self.save_insert_config_now()

    def toggle_phy_chars_and_save(self):
        self.toggle_phy_chars()
        self.save_insert_config_now()

    def choose_insert_input(self):
        p = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
        if p:
            self.insert_input_path = p
            filename = os.path.basename(p)
            self.lbl_insert_in.configure(text=filename, foreground='black')
            self.save_insert_config_now()

    def choose_insert_output(self):
        p = filedialog.askopenfilename(filetypes=[('TXT files', '*.txt')])
        if p:
            self.insert_output_path = p
            filename = os.path.basename(p)
            self.lbl_insert_out.configure(text=filename, foreground='black')
            self.save_insert_config_now()

    def save_insert_config_now(self):
        """Save current insert configuration"""
        if hasattr(self, 'max_chars_var'):
            save_insert_config(
                self.insert_input_path,
                self.insert_output_path,
                self.max_chars_var.get(),
                self.vir_chars_var.get(),
                self.phy_chars_var.get(),
                self.show_vir.get(),
                self.show_phy.get()
            )

    def run_insert(self):
        if not self.insert_input_path or not self.insert_output_path:
            messagebox.showwarning('Input', 'Please choose Excel input and TXT output files.')
            return

        self.pb_insert['value'] = 0
        threading.Thread(target=self.insert_worker, daemon=True).start()

    def insert_worker(self):
        try:
            from openpyxl import load_workbook
            import shutil
            
            # Create backup copy
            backup_path = self.insert_output_path.replace('.txt', '_backup.txt')
            shutil.copy2(self.insert_output_path, backup_path)
            
            # Load Excel data
            wb = load_workbook(self.insert_input_path)
            ws = wb.active
            
            translations = {}
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0] and row[3] and str(row[3]).strip():  # Range, Translate columns, and Translate is not empty
                    range_str = str(row[0])
                    translate_text = str(row[3]).strip()
                    
                    # Skip if translate text is empty after stripping
                    if not translate_text:
                        continue
                    
                    # Check if translate text is "null" - treat as empty string
                    if translate_text.lower() == "null":
                        translate_text = ""
                    
                    # Parse range (e.g., "1069" or "1069-1072")
                    if '-' in range_str:
                        start_num = int(range_str.split('-')[0])
                        end_num = int(range_str.split('-')[1])
                        m_numbers = list(range(start_num, end_num + 1))
                    else:
                        m_numbers = [int(range_str)]
                    
                    # If translate_text is empty (was "null"), set all m_numbers to empty
                    if not translate_text:
                        for m_num in m_numbers:
                            translations[m_num] = ""
                    else:
                        # Apply character replacement
                        processed_text = self.apply_char_replacement(translate_text)
                        
                        # Split text based on max characters
                        split_texts = self.split_text_by_chars(processed_text, self.max_chars_var.get())
                        
                        # Map to m[] numbers
                        for i, m_num in enumerate(m_numbers):
                            if i < len(split_texts):
                                translations[m_num] = split_texts[i]
                            else:
                                # Set remaining m[] numbers to empty string
                                translations[m_num] = ""
            
            # Read original file and replace
            with open(self.insert_output_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            total_lines = len(lines)
            processed_lines = 0
            
            for i, line in enumerate(lines):
                m_match = re_m.match(line)
                if m_match:
                    m_num = int(m_match.group(1))
                    if m_num in translations:
                        # Replace the content without semicolon (uncomment)
                        lines[i] = f'm[{m_num}] = "{translations[m_num]}"\n'
                
                processed_lines += 1
                if processed_lines % 100 == 0:
                    progress = processed_lines / total_lines
                    self.queue.put(('insert_progress', progress))
            
            # Write back to file
            with open(self.insert_output_path, 'w', encoding='utf-8') as f:
                f.writelines(lines)
            
            self.queue.put(('insert_done', f'Insert completed. Backup saved as: {os.path.basename(backup_path)}'))
            
        except Exception as exc:
            self.queue.put(('insert_error', str(exc)))

    def apply_char_replacement(self, text):
        """Replace virtual characters with physical characters"""
        vir_chars = self.vir_chars_var.get()
        phy_chars = self.phy_chars_var.get()
        
        if len(vir_chars) != len(phy_chars):
            return text
        
        result = text
        for vir, phy in zip(vir_chars, phy_chars):
            result = result.replace(vir, phy)
        
        return result

    def split_text_by_chars(self, text, max_chars):
        """Split text by max characters, avoiding breaking words"""
        if len(text) <= max_chars:
            return [text]
        
        result = []
        current = ""
        
        i = 0
        while i < len(text):
            if len(current) + 1 <= max_chars:
                current += text[i]
                i += 1
            else:
                # Find last space or suitable break point
                break_point = len(current)
                for j in range(len(current) - 1, -1, -1):
                    if current[j] in ' 　、。！？':  # Include Japanese punctuation
                        break_point = j + 1
                        break
                
                if break_point == 0:
                    break_point = len(current)
                
                result.append(current[:break_point])
                current = current[break_point:]
                
                if not current and i < len(text):
                    current = text[i]
                    i += 1
        
        if current:
            result.append(current)
        
        return result

    def create_alice_widgets(self):
        # File selection section for Alice tab
        alice_file_frame = ttk.LabelFrame(self.tab_alice, text="File Selection", padding=10)
        alice_file_frame.pack(fill='x', padx=10, pady=5)

        # Ain file selection
        ttk.Button(alice_file_frame, text="Ain file…", width=12, command=self.choose_ain_file)\
            .grid(row=0, column=0, padx=5, sticky='w')
        self.lbl_ain_file = ttk.Label(alice_file_frame, text='–', foreground='gray')
        self.lbl_ain_file.grid(row=0, column=1, sticky='w', padx=(10, 0))

        # TXT file selection
        ttk.Button(alice_file_frame, text="TXT file…", width=12, command=self.choose_txt_file)\
            .grid(row=1, column=0, padx=5, pady=6, sticky='w')
        self.lbl_txt_file = ttk.Label(alice_file_frame, text='–', foreground='gray')
        self.lbl_txt_file.grid(row=1, column=1, sticky='w', padx=(10, 0))

        # Output path selection
        ttk.Button(alice_file_frame, text="Output path…", width=12, command=self.choose_output_ain)\
            .grid(row=2, column=0, padx=5, pady=6, sticky='w')
        self.lbl_output_ain = ttk.Label(alice_file_frame, text='–', foreground='gray')
        self.lbl_output_ain.grid(row=2, column=1, sticky='w', padx=(10, 0))

        # Pack section
        alice_action_frame = ttk.Frame(self.tab_alice)
        alice_action_frame.pack(fill='x', padx=10, pady=20)

        # Pack button
        pack_btn = ttk.Button(alice_action_frame, text="Pack Ain File", width=25, command=self.run_alice_pack)
        pack_btn.pack(pady=10)

        # Progress for alice pack
        ttk.Label(alice_action_frame, text="Pack Progress:").pack(anchor='w')
        self.pb_alice = ttk.Progressbar(alice_action_frame, orient='horizontal', length=500, mode='indeterminate')
        self.pb_alice.pack(pady=(5, 0))

        # Initialize paths
        self.ain_file_path = ''
        self.txt_file_path = ''
        self.output_ain_path = ''

    def choose_ain_file(self):
        p = filedialog.askopenfilename(filetypes=[('AIN files', '*.ain')])
        if p:
            self.ain_file_path = p
            filename = os.path.basename(p)
            self.lbl_ain_file.configure(text=filename, foreground='black')
            self.save_alice_config_now()

    def choose_txt_file(self):
        p = filedialog.askopenfilename(filetypes=[('TXT files', '*.txt')])
        if p:
            self.txt_file_path = p
            filename = os.path.basename(p)
            self.lbl_txt_file.configure(text=filename, foreground='black')
            self.save_alice_config_now()

    def choose_output_ain(self):
        p = filedialog.asksaveasfilename(defaultextension='.ain', filetypes=[('AIN files', '*.ain')])
        if p:
            self.output_ain_path = p
            filename = os.path.basename(p)
            self.lbl_output_ain.configure(text=filename, foreground='black')
            self.save_alice_config_now()

    def save_alice_config_now(self):
        """Save current alice configuration"""
        # Load existing config first
        existing_config = load_config() or {}
        
        alice_config = {
            'ain_file_path': self.ain_file_path,
            'txt_file_path': self.txt_file_path,
            'output_ain_path': self.output_ain_path
        }
        
        existing_config['alice_config'] = alice_config
        
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(existing_config, f, ensure_ascii=False, indent=2)

    def run_alice_pack(self):
        if not self.ain_file_path or not self.txt_file_path or not self.output_ain_path:
            messagebox.showwarning('Input', 'Please choose all required files: Ain file, TXT file, and Output path.')
            return

        self.pb_alice.start()
        threading.Thread(target=self.alice_pack_worker, daemon=True).start()

    def alice_pack_worker(self):
        try:
            import subprocess
            
            # Get alice.exe path (in alice-tool folder next to this script)
            alice_exe = get_resource_path(os.path.join('alice-tool', 'alice.exe'))
            
            if not os.path.exists(alice_exe):
                # Fallback: try relative to script directory
                script_dir = os.path.dirname(os.path.abspath(__file__))
                alice_exe = os.path.join(script_dir, 'alice-tool', 'alice.exe')
                
                if not os.path.exists(alice_exe):
                    self.queue.put(('alice_error', f'alice.exe not found at: {alice_exe}'))
                    return
            
            # Build command: alice ain edit -t out.txt -o Tsumamigui3.ain Tsumamigui.ain
            cmd = [
                alice_exe,
                'ain', 'edit',
                '-t', self.txt_file_path,
                '-o', self.output_ain_path,
                self.ain_file_path
            ]
            
            # Run command
            result = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
            
            if result.returncode == 0:
                self.queue.put(('alice_done', f'Pack completed successfully!\nOutput: {os.path.basename(self.output_ain_path)}'))
            else:
                error_msg = result.stderr if result.stderr else result.stdout
                self.queue.put(('alice_error', f'Pack failed:\n{error_msg}'))
                
        except Exception as exc:
            self.queue.put(('alice_error', str(exc)))


if __name__ == '__main__':
    App().mainloop()